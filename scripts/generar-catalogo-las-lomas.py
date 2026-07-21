import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


REQUIRED_COLUMNS = {
    "Path SVG",
    "Manzana",
    "Lote",
    "Area",
    "Precio",
    "Estado",
}

EXPECTED_LOT_COUNT = 213
ALLOWED_STATES = {"DISPONIBLE", "SEPARADO", "VENDIDO"}


def sql_text(value):
    return "'" + str(value).replace("'", "''") + "'"


def number_sql(value):
    return f"{float(value):.2f}"


def read_catalog(workbook_path):
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    sheet = workbook["SUPABASE"]
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]

    missing = REQUIRED_COLUMNS.difference(headers)
    if missing:
        raise ValueError(f"Faltan columnas obligatorias: {sorted(missing)}")

    index = {name: headers.index(name) for name in REQUIRED_COLUMNS}
    catalog = []
    for row in rows[1:]:
        if not row[index["Path SVG"]]:
            continue
        catalog.append(
            {
                "mz": str(row[index["Manzana"]]).strip().upper(),
                "lote": int(row[index["Lote"]]),
                "area": round(float(row[index["Area"]]), 2),
                "precio": round(float(row[index["Precio"]]), 2),
                "estado": str(row[index["Estado"]]).strip().upper(),
                "svg_id": str(row[index["Path SVG"]]).strip(),
            }
        )
    return catalog


def validate(catalog, svg_path):
    lot_keys = [(item["mz"], item["lote"]) for item in catalog]
    path_ids = [item["svg_id"] for item in catalog]
    if len(catalog) != EXPECTED_LOT_COUNT:
        raise ValueError(
            f"Se esperaban {EXPECTED_LOT_COUNT} lotes y se encontraron {len(catalog)}"
        )
    if len(set(lot_keys)) != len(lot_keys):
        raise ValueError("El Excel contiene manzana/lote duplicados")
    if len(set(path_ids)) != len(path_ids):
        raise ValueError("El Excel contiene Path SVG duplicados")
    invalid_rows = [
        item
        for item in catalog
        if not item["mz"]
        or item["lote"] <= 0
        or item["area"] <= 0
        or item["precio"] < 0
        or item["estado"] not in ALLOWED_STATES
        or not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_.:-]*", item["svg_id"])
    ]
    if invalid_rows:
        raise ValueError(f"El Excel contiene filas invalidas: {invalid_rows}")

    svg_ids = set(re.findall(r'\bid=["\']([^"\']+)["\']', svg_path.read_text(encoding="utf-8")))
    missing = sorted(set(path_ids).difference(svg_ids))
    if missing:
        raise ValueError(f"Hay Path SVG que no existen en el plano: {missing}")


def assign_stable_ids(catalog, previous_json):
    previous = json.loads(previous_json.read_text(encoding="utf-8"))
    ids_by_lot = {(str(item["mz"]).upper(), int(item["lote"])): int(item["id"]) for item in previous}
    next_id = max(ids_by_lot.values(), default=0) + 1

    for item in catalog:
        key = (item["mz"], item["lote"])
        item["id"] = ids_by_lot.get(key)
        if item["id"] is None:
            item["id"] = next_id
            next_id += 1

    catalog.sort(key=lambda item: item["id"])


def build_migration(catalog):
    values = []
    for item in catalog:
        values.append(
            "  ("
            + ", ".join(
                [
                    str(item["id"]),
                    sql_text(item["mz"]),
                    str(item["lote"]),
                    number_sql(item["area"]),
                    number_sql(item["precio"]),
                    sql_text(item["estado"]),
                    sql_text(item["svg_id"]),
                ]
            )
            + ")"
        )

    return f"""-- 016_actualizar_catalogo_213_lotes.sql
-- Generado desde supabase-las-lomas-nuevo.xlsx. No editar filas manualmente.
-- Conserva el id de los lotes existentes para no romper relaciones del CRM.

begin;

do $$
begin
  if exists (select 1 from public.clientes)
     or exists (select 1 from public.separaciones)
     or exists (select 1 from public.historial_lotes)
     or exists (select 1 from public.leads_publicos)
     or exists (select 1 from public.meta_lead_events)
     or exists (select 1 from public.seguimientos_clientes)
     or exists (select 1 from public.metas_comerciales)
     or exists (select 1 from public.cotizaciones)
     or exists (select 1 from public.separacion_expedientes)
     or exists (select 1 from public.separacion_documentos)
  then
    raise exception 'La migracion del catalogo requiere los datos operativos limpios. Ejecuta primero 015_crm_preparar_produccion.sql.';
  end if;
end $$;

create temporary table nuevo_catalogo_las_lomas (
  id bigint primary key,
  mz text not null,
  lote integer not null,
  area numeric not null,
  precio numeric not null,
  estado text not null,
  svg_id text not null unique,
  unique (mz, lote)
) on commit drop;

insert into nuevo_catalogo_las_lomas
  (id, mz, lote, area, precio, estado, svg_id)
values
{',\n'.join(values)};

-- Los datos operativos ya estan vacios. Sustituir el catalogo completo evita
-- colisiones entre svg_id antiguos y nuevos durante una actualizacion parcial.
delete from public.las_lomas_lotes;

insert into public.las_lomas_lotes
  (id, mz, lote, area, precio, estado, svg_id, cliente_id, asesor_id, updated_at)
select
  id, mz, lote, area, precio, estado, svg_id, null, null, now()
from nuevo_catalogo_las_lomas
on conflict (id) do update
set mz = excluded.mz,
    lote = excluded.lote,
    area = excluded.area,
    precio = excluded.precio,
    estado = excluded.estado,
    svg_id = excluded.svg_id,
    cliente_id = null,
    asesor_id = null,
    updated_at = now();

do $$
declare
  total bigint;
begin
  select count(*) into total from public.las_lomas_lotes;
  if total <> 213 then
    raise exception 'Catalogo incompleto: se esperaban 213 lotes y quedaron %', total;
  end if;
end $$;

select setval(
  pg_get_serial_sequence('public.las_lomas_lotes', 'id'),
  (select max(id) from public.las_lomas_lotes),
  true
)
where pg_get_serial_sequence('public.las_lomas_lotes', 'id') is not null;

commit;

select count(*)::bigint as total_lotes from public.las_lomas_lotes;
select mz, count(*)::bigint as lotes
from public.las_lomas_lotes
group by mz
order by mz;
"""


def main():
    parser = argparse.ArgumentParser(description="Genera el catalogo de 213 lotes y su migracion SQL")
    parser.add_argument("--xlsx", required=True, type=Path)
    parser.add_argument("--svg", required=True, type=Path)
    parser.add_argument("--previous-json", required=True, type=Path)
    parser.add_argument("--output-json", required=True, type=Path)
    parser.add_argument("--output-sql", required=True, type=Path)
    args = parser.parse_args()

    catalog = read_catalog(args.xlsx)
    validate(catalog, args.svg)
    assign_stable_ids(catalog, args.previous_json)

    args.output_json.write_text(
        json.dumps(catalog, ensure_ascii=True, indent=2) + "\n",
        encoding="utf-8",
    )
    args.output_sql.write_text(build_migration(catalog), encoding="utf-8")
    print(f"Catalogo generado: {len(catalog)} lotes")


if __name__ == "__main__":
    main()
